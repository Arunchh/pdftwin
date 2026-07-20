import tempfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import fitz
import pdfplumber
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from pdf2docx import Converter

from utils import normalize_unicode

MAX_CELL_IMAGE_WIDTH = 140
MAX_CELL_IMAGE_HEIGHT = 140
ROW_CLUSTER_TOLERANCE = 12
MIN_EXTRACTABLE_TEXT = 40


@dataclass
class PageImage:
    bbox: tuple[float, float, float, float]
    data: bytes


def _sanitize_cell(value) -> str:
    return normalize_unicode("" if value is None else str(value))


def _normalize_row(row: list) -> tuple[str, ...]:
    return tuple(_sanitize_cell(cell).strip().lower() for cell in row)


def _rows_match(left: list, right: list) -> bool:
    if not left or not right or len(left) != len(right):
        return False
    return _normalize_row(left) == _normalize_row(right)


def _bbox_center(bbox: tuple[float, float, float, float]) -> tuple[float, float]:
    return ((bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2)


def _center_in_bbox(
    outer: tuple[float, float, float, float],
    inner: tuple[float, float, float, float],
) -> bool:
    cx, cy = _bbox_center(inner)
    return outer[0] <= cx <= outer[2] and outer[1] <= cy <= outer[3]


def _scale_image(image: XLImage, max_width: int, max_height: int) -> None:
    if image.width <= 0 or image.height <= 0:
        return
    scale = min(max_width / image.width, max_height / image.height, 1.0)
    if scale < 1.0:
        image.width = int(image.width * scale)
        image.height = int(image.height * scale)


def _get_page_images(page: fitz.Page) -> list[PageImage]:
    images: list[PageImage] = []
    for block in page.get_text("dict", sort=True)["blocks"]:
        if block.get("type") != 1:
            continue
        image_bytes = block.get("image")
        if not image_bytes:
            continue
        images.append(PageImage(tuple(block["bbox"]), image_bytes))
    return images


def _find_image_for_cell(
    images: list[PageImage],
    cell_bbox: tuple[float, float, float, float] | None,
    used: set[int],
) -> PageImage | None:
    if not cell_bbox:
        return None

    for index, image in enumerate(images):
        if index in used:
            continue
        if _center_in_bbox(cell_bbox, image.bbox):
            used.add(index)
            return image
    return None


def _insert_cell_image(sheet, row: int, col: int, image_bytes: bytes) -> None:
    image = XLImage(BytesIO(image_bytes))
    _scale_image(image, MAX_CELL_IMAGE_WIDTH, MAX_CELL_IMAGE_HEIGHT)
    anchor = f"{get_column_letter(col)}{row}"
    sheet.add_image(image, anchor)

    column = get_column_letter(col)
    current_width = sheet.column_dimensions[column].width or 8
    sheet.column_dimensions[column].width = max(current_width, image.width / 7)

    current_height = sheet.row_dimensions[row].height or 15
    sheet.row_dimensions[row].height = max(current_height, image.height * 0.75)


def _write_text_cell(sheet, row: int, col: int, value: str) -> None:
    cleaned = _sanitize_cell(value)
    if cleaned.strip():
        sheet.cell(row=row, column=col, value=cleaned)


def _write_data_row(
    sheet,
    row: int,
    values: list[str],
    cell_bboxes: list[tuple[float, float, float, float] | None],
    page_images: list[PageImage],
    used_images: set[int],
) -> None:
    for col_index, value in enumerate(values, start=1):
        cell_bbox = cell_bboxes[col_index - 1] if col_index - 1 < len(cell_bboxes) else None
        image = _find_image_for_cell(page_images, cell_bbox, used_images)
        text = _sanitize_cell(value)

        if image and len(text.strip()) <= 2:
            _insert_cell_image(sheet, row, col_index, image.data)
        else:
            _write_text_cell(sheet, row, col_index, text)
            if image:
                _insert_cell_image(sheet, row, col_index, image.data)


def _cluster_layout_items(page: fitz.Page) -> list[list[dict]]:
    items: list[dict] = []

    for block in page.get_text("dict", sort=True)["blocks"]:
        bbox = tuple(block["bbox"])
        if block.get("type") == 0:
            text = "".join(
                span["text"]
                for line in block.get("lines", [])
                for span in line.get("spans", [])
            ).strip()
            if text:
                items.append({"kind": "text", "bbox": bbox, "text": _sanitize_cell(text)})
        elif block.get("type") == 1 and block.get("image"):
            items.append({"kind": "image", "bbox": bbox, "data": block["image"]})

    if not items:
        return []

    items.sort(key=lambda item: _bbox_center(item["bbox"])[1])
    rows: list[list[dict]] = [[items[0]]]
    current_y = _bbox_center(items[0]["bbox"])[1]

    for item in items[1:]:
        y = _bbox_center(item["bbox"])[1]
        if abs(y - current_y) <= ROW_CLUSTER_TOLERANCE:
            rows[-1].append(item)
        else:
            rows.append([item])
            current_y = y

    return [sorted(row_items, key=lambda item: item["bbox"][0]) for row_items in rows]


def _layout_rows_to_values(layout_rows: list[list[dict]]) -> list[list[str]]:
    values: list[list[str]] = []
    for row_items in layout_rows:
        row_values: list[str] = []
        for item in row_items:
            if item["kind"] == "text":
                row_values.append(item["text"])
            else:
                row_values.append("")
        if any(cell.strip() for cell in row_values) or any(
            item["kind"] == "image" for item in row_items
        ):
            values.append(row_values)
    return values


def _write_layout_page(
    sheet,
    page: fitz.Page,
    start_row: int,
    known_header: list[str] | None,
) -> tuple[int, list[str] | None]:
    layout_rows = _cluster_layout_items(page)
    if not layout_rows:
        return start_row, known_header

    page_images = _get_page_images(page)
    used_images: set[int] = set()
    row = start_row
    header = known_header

    for row_items in layout_rows:
        values = []
        bboxes = []
        for item in row_items:
            if item["kind"] == "text":
                values.append(item["text"])
            else:
                values.append("")
            bboxes.append(item["bbox"])

        if header and _rows_match(values, header):
            continue

        if header is None and any(value.strip() for value in values):
            header = values

        _write_data_row(sheet, row, values, bboxes, page_images, used_images)
        row += 1

    return row, header


def _write_pymupdf_table(
    sheet,
    table,
    page_images: list[PageImage],
    start_row: int,
    known_header: list[str] | None,
) -> tuple[int, list[str] | None]:
    data = table.extract() or []
    if not data:
        return start_row, known_header

    used_images: set[int] = set()
    row = start_row
    header = known_header

    for row_index, values in enumerate(data):
        if header and row_index == 0 and _rows_match(values, header):
            continue
        if header is None and row_index == 0:
            header = [_sanitize_cell(value) for value in values]

        cell_bboxes = []
        if row_index < len(table.rows):
            cell_bboxes = list(table.rows[row_index].cells)

        _write_data_row(sheet, row, values, cell_bboxes, page_images, used_images)
        row += 1

    return row, header


def _write_pdfplumber_tables(
    sheet,
    page,
    start_row: int,
    known_header: list[str] | None,
) -> tuple[int, list[str] | None]:
    tables = [table for table in (page.extract_tables() or []) if table]
    if not tables:
        return start_row, known_header

    row = start_row
    header = known_header

    for table in tables:
        for row_index, values in enumerate(table):
            cleaned = [_sanitize_cell(value) for value in values]
            if header and row_index == 0 and _rows_match(cleaned, header):
                continue
            if header is None and row_index == 0:
                header = cleaned

            for col_index, value in enumerate(cleaned, start=1):
                _write_text_cell(sheet, row, col_index, value)
            row += 1

    return row, header


def _count_extractable_content(doc: fitz.Document) -> tuple[int, int]:
    text_chars = 0
    table_count = 0

    for page in doc:
        text_chars += len(page.get_text().strip())
        table_count += len(page.find_tables().tables)

    return text_chars, table_count


def pdf_to_word(content: bytes) -> bytes:
    with tempfile.TemporaryDirectory() as tmp_dir:
        pdf_path = Path(tmp_dir) / "input.pdf"
        docx_path = Path(tmp_dir) / "output.docx"

        pdf_path.write_bytes(content)

        converter = Converter(str(pdf_path))
        try:
            converter.convert(str(docx_path))
        finally:
            converter.close()

        return docx_path.read_bytes()


def pdf_to_excel(content: bytes) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    doc = fitz.open(stream=content, filetype="pdf")
    try:
        if doc.is_encrypted:
            raise ValueError(
                "This PDF is password-protected. Unlock it first using the Lock / Unlock tab."
            )

        if not doc.page_count:
            sheet.cell(row=1, column=1, value="No pages found in PDF.")
        else:
            text_chars, table_count = _count_extractable_content(doc)
            if text_chars < MIN_EXTRACTABLE_TEXT and table_count == 0:
                raise ValueError(
                    "This PDF looks like a scan or image-only file. PDF to Excel works best "
                    "with selectable text and clear rows/columns (like reports, catalogs, or "
                    "invoices). Re-export the PDF with real text, or use a dedicated OCR tool "
                    "for scanned documents."
                )

            current_row = 1
            known_header: list[str] | None = None
            wrote_anything = False

            with pdfplumber.open(BytesIO(content)) as pdf:
                for page_index in range(doc.page_count):
                    page = doc[page_index]
                    page_images = _get_page_images(page)
                    tables = page.find_tables().tables

                    if tables:
                        for table in tables:
                            current_row, known_header = _write_pymupdf_table(
                                sheet,
                                table,
                                page_images,
                                current_row,
                                known_header,
                            )
                            wrote_anything = True
                        continue

                    plumber_page = (
                        pdf.pages[page_index] if page_index < len(pdf.pages) else None
                    )
                    if plumber_page:
                        plumber_tables = [
                            table for table in (plumber_page.extract_tables() or []) if table
                        ]
                        if plumber_tables:
                            current_row, known_header = _write_pdfplumber_tables(
                                sheet,
                                plumber_page,
                                current_row,
                                known_header,
                            )
                            wrote_anything = True
                            continue

                    next_row, known_header = _write_layout_page(
                        sheet,
                        page,
                        current_row,
                        known_header,
                    )
                    if next_row > current_row:
                        wrote_anything = True
                    current_row = next_row

            if not wrote_anything:
                sheet.cell(
                    row=1,
                    column=1,
                    value=(
                        "No structured rows could be extracted. Make sure the PDF has "
                        "selectable text arranged in rows/columns."
                    ),
                )

            for column_index in range(1, 16):
                column = get_column_letter(column_index)
                current_width = sheet.column_dimensions[column].width
                if not current_width or current_width < 12:
                    sheet.column_dimensions[column].width = 12
    finally:
        doc.close()

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
